import customtkinter as ctk
import base64
import ctypes
import json
import os
import re
import sys
import threading
import time
import traceback
from datetime import datetime
from tkinter import filedialog, messagebox
from ctypes import wintypes

import openpyxl
from PIL import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from selenium import webdriver
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


LOGIN_URL = "https://motos.man.eu/motos/frame/frame_en.html"
TARGET_CODE = "20"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEBUG_DIR = os.path.join(BASE_DIR, "debug")
ENABLE_DEBUG_DETAILS = False
LOG_STATUS_TO_FILE = False
CONFIG_DIR = os.path.join(os.environ.get("APPDATA", BASE_DIR), "MAN_Ekstraktor")
CONFIG_FILE = os.path.join(CONFIG_DIR, "config.json")
LOG_FILE = os.path.join(CONFIG_DIR, "app.log")
BANNER_FILE = os.path.join("assets", "man_motos_excel_banner.png")
ICON_FILE = os.path.join("assets", "ex_man.ico")

COLUMNS = [
    "NAZWA / SKROT", "NUMER TELEFONU", "STAWKA", "FAKTURA GOTOWKA",
    "ZALADUNEK PLANOWANA DATA", "ZALADUNEK KRAJ", "ZALADUNEK KOD POCZTOWY",
    "ZALADUNEK MIASTO", "ZALADUNEK ULICA NUMER",
    "DOSTAWA PLANOWANA DATA", "DOSTAWA KRAJ", "DOSTAWA KOD POCZTOWY",
    "DOSTAWA MIASTO", "DOSTAWA ULICA NUMER",
    "VIN NUMER REFERENCYJNY", "TOWAR NAZWA URZADZENIA",
    "DLUGOSC [mm]", "SZEROKOSC [mm]", "WYSOKOSC [mm]",
    "ROZSTAW OSI [mm]", "WAGA [kg]",
    "BOOKING PRZESYLKA TMS", "WYSZUKIWANIE",
    "KIEROWCA ZALADUNEK", "FORMATKA DLA KIEROWCY",
]

APP_VERSION = "1.1"

LIST_UNITS_JS = r"""
return (function() {
  function collect(win, path, units) {
    try {
      const doc = win.document;
      Array.from(doc.querySelectorAll('tr')).forEach(function(row) {
        const cells = Array.from(row.querySelectorAll('td')).map(function(cell) {
          return (cell.innerText || '').trim().replace(/\u00a0/g, ' ');
        });
        const nonEmpty = cells.filter(Boolean);
        if (nonEmpty[0] !== '20') return;
        if (!/^[A-Z0-9]{6,8}$/.test(nonEmpty[1] || '')) return;

        const link = row.querySelector("a[href*='transportDetails'], a[href*='transport']");
        units.push({
          frame: path,
          vehicleNo: nonEmpty[1] || '',
          pickupCode: nonEmpty[2] || '',
          pickupDate: nonEmpty[3] || '',
          country: nonEmpty[4] || '',
          dest: nonEmpty[5] || '',
          toNo: nonEmpty[16] || '',
          href: link ? link.href : '',
          cells: nonEmpty
        });
      });

      for (let i = 0; i < win.frames.length; i++) {
        collect(win.frames[i], path + '/frame[' + i + ']', units);
      }
    } catch(e) {}
  }

  const units = [];
  collect(window, 'top', units);
  return JSON.stringify({units: units, unitCount: units.length});
})();
"""

CLICK_LABEL_JS = r"""
const label = arguments[0];
return (function clickIn(win) {
  try {
    const doc = win.document;
    const candidates = Array.from(doc.querySelectorAll('a, span, td, div, input, button'));
    for (const el of candidates) {
      const text = ((el.innerText || el.value || '') + '').trim().replace(/\u00a0/g, ' ');
      if (text !== label) continue;
      const clickable = el.closest('a') || el;
      clickable.click();
      return true;
    }
    for (let i = 0; i < win.frames.length; i++) {
      if (clickIn(win.frames[i])) return true;
    }
  } catch (e) {}
  return false;
})(window);
"""

DETAILS_EXTRACT_JS = r"""
return (function() {
  function collect(win, out) {
    try {
      const doc = win.document;
      const bodyText = (doc.body && doc.body.innerText) ? doc.body.innerText : '';
      out.push(bodyText);
      for (let i = 0; i < win.frames.length; i++) collect(win.frames[i], out);
    } catch (e) {}
  }

  function clean(x) {
    return (x || '').trim().replace(/\u00a0/g, ' ');
  }

  function linesOf(text) {
    return text.split(/\r?\n/).map(clean).filter(Boolean);
  }

  function valueAfterLabel(text, labels) {
    const lines = linesOf(text);
    for (let i = 0; i < lines.length; i++) {
      for (const label of labels) {
        if (lines[i] === label && i + 1 < lines.length) return lines[i + 1];
        if (lines[i].startsWith(label)) {
          const value = clean(lines[i].slice(label.length));
          if (value) return value;
        }
      }
    }
    return '';
  }

  function splitAddressLine(line) {
    const parts = clean(line).split(/\s+/).filter(Boolean);
    if (parts.length >= 3) {
      return {country: parts[0] || '', zipcode: parts[1] || '', city: parts.slice(2).join(' ')};
    }
    if (parts.length === 2) {
      return {country: parts[0] || '', zipcode: '', city: parts[1] || ''};
    }
    return {country: '', zipcode: '', city: line || ''};
  }

  function parseAddressBlock(lines) {
    const last = splitAddressLine(lines[2] || '');
    return {
      name: lines[0] || '',
      street: lines[1] || '',
      country: last.country,
      zipcode: last.zipcode,
      city: last.city
    };
  }

  function extractAddresses(text) {
    const lines = linesOf(text);
    const result = {pickup: {}, delivery: {}};

    for (let i = 0; i < lines.length; i++) {
      if (lines[i] !== 'Pick up:') continue;
      if (lines[i + 2] !== 'Delivery:') continue;

      const start = i + 4;
      const pickupLines = lines.slice(start, start + 3);
      const deliveryLines = lines.slice(start + 3, start + 6);

      if (pickupLines.length === 3 && deliveryLines.length === 3) {
        result.pickup = parseAddressBlock(pickupLines);
        result.delivery = parseAddressBlock(deliveryLines);
        return result;
      }
    }

    return result;
  }

  function extractTelephone(text) {
    const lines = linesOf(text);
    const numbers = [];
    for (let i = 0; i < lines.length; i++) {
      if (lines[i] === 'Telephone number:' && i + 1 < lines.length) {
        const value = lines[i + 1];
        if (value && value !== 'Fax number:') numbers.push(value);
      }
    }
    return numbers.length ? numbers[numbers.length - 1] : '';
  }

  const chunks = [];
  collect(window, chunks);
  const text = chunks.join('\n');
  const addresses = extractAddresses(text);

  return JSON.stringify({
    typeCab: valueAfterLabel(text, ['Type / cab:', 'Type / cab']),
    vin: valueAfterLabel(text, ['VIN:', 'VIN']),
    totalLength: valueAfterLabel(text, ['Total length:', 'Total length']),
    height: valueAfterLabel(text, ['Height:', 'Height']),
    weight: valueAfterLabel(text, ['Weight:', 'Weight']),
    wheelOverhang: valueAfterLabel(text, ['Wheel. overhang:', 'Wheel. overhang']),
    vehicleType: valueAfterLabel(text, ['Vehicle type:', 'Vehicle type']),
    pickupUntil: valueAfterLabel(text, ['Pick up until:', 'Pick up untill:', 'Pick up until', 'Pick up untill']),
    latestDelivery: valueAfterLabel(text, ['Latest delivery:', 'Latest delivery']),
    telephone: extractTelephone(text),
    pickupAddress: addresses.pickup,
    deliveryAddress: addresses.delivery
  });
})();
"""


class FileLockedError(Exception):
    def __init__(self, filepath):
        super().__init__(filepath)
        self.filepath = filepath


def resource_path(relative_path):
    base_path = getattr(sys, "_MEIPASS", BASE_DIR)
    return os.path.join(base_path, relative_path)


def write_app_log(message, force=False):
    if not force and not LOG_STATUS_TO_FILE:
        return

    try:
        os.makedirs(CONFIG_DIR, exist_ok=True)
        with open(LOG_FILE, "a", encoding="utf-8") as handle:
            handle.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  {message}\n")
    except Exception:
        pass


def write_exception_log(context, exc):
    write_app_log(f"{context}: {exc}", force=True)
    try:
        os.makedirs(CONFIG_DIR, exist_ok=True)
        with open(LOG_FILE, "a", encoding="utf-8") as handle:
            handle.write(traceback.format_exc())
            handle.write("\n")
    except Exception:
        pass


class DATA_BLOB(ctypes.Structure):
    _fields_ = [
        ("cbData", wintypes.DWORD),
        ("pbData", ctypes.POINTER(ctypes.c_char)),
    ]


def encrypt_for_current_user(text):
    raw = text.encode("utf-8")
    in_buffer = ctypes.create_string_buffer(raw)
    in_blob = DATA_BLOB(len(raw), ctypes.cast(in_buffer, ctypes.POINTER(ctypes.c_char)))
    out_blob = DATA_BLOB()

    ok = ctypes.windll.crypt32.CryptProtectData(
        ctypes.byref(in_blob),
        None,
        None,
        None,
        None,
        0,
        ctypes.byref(out_blob),
    )

    if not ok:
        raise ctypes.WinError()

    try:
        encrypted = ctypes.string_at(out_blob.pbData, out_blob.cbData)
        return base64.b64encode(encrypted).decode("ascii")
    finally:
        ctypes.windll.kernel32.LocalFree(out_blob.pbData)


def decrypt_for_current_user(encoded_text):
    encrypted = base64.b64decode(encoded_text.encode("ascii"))
    in_buffer = ctypes.create_string_buffer(encrypted)
    in_blob = DATA_BLOB(len(encrypted), ctypes.cast(in_buffer, ctypes.POINTER(ctypes.c_char)))
    out_blob = DATA_BLOB()

    ok = ctypes.windll.crypt32.CryptUnprotectData(
        ctypes.byref(in_blob),
        None,
        None,
        None,
        None,
        0,
        ctypes.byref(out_blob),
    )

    if not ok:
        raise ctypes.WinError()

    try:
        decrypted = ctypes.string_at(out_blob.pbData, out_blob.cbData)
        return decrypted.decode("utf-8")
    finally:
        ctypes.windll.kernel32.LocalFree(out_blob.pbData)


def load_saved_login():
    if not os.path.exists(CONFIG_FILE):
        return {"remember": False, "email": "", "password": ""}

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as handle:
            config = json.load(handle)

        if not config.get("remember"):
            return {"remember": False, "email": "", "password": ""}

        return {
            "remember": True,
            "email": clean_text(config.get("email", "")),
            "password": decrypt_for_current_user(config.get("password", "")),
        }
    except Exception:
        return {"remember": False, "email": "", "password": ""}


def save_login(login, password):
    os.makedirs(CONFIG_DIR, exist_ok=True)
    config = {
        "remember": True,
        "email": login,
        "password": encrypt_for_current_user(password),
    }

    with open(CONFIG_FILE, "w", encoding="utf-8") as handle:
        json.dump(config, handle, ensure_ascii=False, indent=2)


def delete_saved_login():
    try:
        if os.path.exists(CONFIG_FILE):
            os.remove(CONFIG_FILE)
    except Exception:
        pass


def clean_text(txt):
    return (txt or "").replace("\u00a0", " ").strip()


def following_text(driver, label):
    label = clean_text(label)
    xpaths = [
        f"//td[normalize-space(translate(., '\u00a0', ' '))='{label}']/following-sibling::td[1]",
        f"//*[normalize-space(translate(., '\u00a0', ' '))='{label}']/following::td[1]",
        f"//td[contains(normalize-space(translate(., '\u00a0', ' ')), '{label}')]/following-sibling::td[1]",
        f"//*[contains(normalize-space(translate(., '\u00a0', ' ')), '{label}')]/following::td[1]",
    ]

    for xp in xpaths:
        try:
            els = driver.find_elements(By.XPATH, xp)
            for el in els:
                txt = clean_text(el.text)
                if txt:
                    return txt
        except Exception:
            continue

    return ""


def following_text_any(driver, labels):
    for label in labels:
        value = following_text(driver, label)
        if value:
            return value
    return ""


def visible_lines(driver):
    try:
        text = driver.find_element(By.TAG_NAME, "body").text
    except Exception:
        return []

    return [clean_text(line) for line in text.splitlines() if clean_text(line)]


def value_after_label_in_text(driver, labels):
    labels = [clean_text(label) for label in labels]
    lines = visible_lines(driver)

    for idx, line in enumerate(lines):
        for label in labels:
            if line == label and idx + 1 < len(lines):
                return lines[idx + 1]

            if line.startswith(label):
                value = clean_text(line[len(label):])
                if value:
                    return value

    return ""


def field_value(driver, labels):
    value = following_text_any(driver, labels)
    if value:
        return value

    return value_after_label_in_text(driver, labels)


def is_bialezyce_100(street):
    normalized = clean_text(street).lower()
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized == "bialezyce 100"


def append_after_comma(base_value, extra_value):
    base_value = clean_text(base_value)
    extra_value = clean_text(extra_value)

    if not extra_value:
        return base_value

    if not base_value:
        return extra_value

    parts = [clean_text(part) for part in base_value.split(",")]
    if extra_value in parts:
        return base_value

    return f"{base_value}, {extra_value}"


def parse_address(block):
    lines = [clean_text(line) for line in block.splitlines() if clean_text(line)]

    firm = ""
    street = ""
    country = ""
    zipcode = ""
    city = ""

    if len(lines) >= 1:
        firm = lines[0]
    if len(lines) >= 2:
        street = lines[1]
    if len(lines) >= 3:
        parts = lines[-1].split()
        if len(parts) >= 3:
            country = parts[0]
            zipcode = parts[1]
            city = " ".join(parts[2:])
        elif len(parts) == 2:
            country = parts[0]
            city = parts[1]
        elif len(parts) == 1:
            city = parts[0]

    return firm, street, country, zipcode, city


def looks_like_address_block(text):
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
    if len(lines) < 3:
        return False

    return bool(re.match(r"^[A-Z]{2}\s+\S+\s+.+", lines[-1]))


def address_block_after_label(driver, label):
    label = clean_text(label)
    xpaths = [
        f"//*[normalize-space(translate(., '\u00a0', ' '))='{label}']/following::td",
    ]

    for xp in xpaths:
        try:
            cells = driver.find_elements(By.XPATH, xp)
            for cell in cells:
                txt = clean_text(cell.text)
                if not txt:
                    continue
                if txt in ["Pick up:", "Delivery:", "Telephone number:", "Fax number:"]:
                    continue
                if looks_like_address_block(txt):
                    return txt
        except Exception:
            continue

    return ""


def address_block_near_label(driver, label):
    label = clean_text(label)

    try:
        label_elements = driver.find_elements(
            By.XPATH,
            f"//*[normalize-space(translate(., '\u00a0', ' '))='{label}']",
        )
        candidate_cells = driver.find_elements(By.XPATH, "//td|//div|//span")
    except Exception:
        return ""

    candidates = []
    for cell in candidate_cells:
        try:
            txt = clean_text(cell.text)
            if not looks_like_address_block(txt):
                continue

            rect = cell.rect
            if rect.get("width", 0) <= 0 or rect.get("height", 0) <= 0:
                continue

            candidates.append((cell, txt, rect))
        except Exception:
            continue

    best = None

    for label_el in label_elements:
        try:
            label_rect = label_el.rect
            label_x = label_rect.get("x", 0)
            label_y = label_rect.get("y", 0)
            label_w = label_rect.get("width", 0)
            label_center_x = label_x + label_w / 2
        except Exception:
            continue

        for _, txt, rect in candidates:
            cand_x = rect.get("x", 0)
            cand_y = rect.get("y", 0)
            cand_w = rect.get("width", 0)
            cand_center_x = cand_x + cand_w / 2

            # W TO details adres jest po prawej stronie etykiety Pick up/Delivery.
            if cand_center_x < label_center_x:
                continue

            dy = abs(cand_y - label_y)
            dx = cand_x - label_x
            score = dy * 10 + abs(dx)

            if best is None or score < best[0]:
                best = (score, txt)

    if best:
        return best[1]

    return ""


def address_block_from_text(driver, label):
    lines = visible_lines(driver)

    for idx, line in enumerate(lines):
        if line != label:
            continue

        collected = []
        for item in lines[idx + 1:]:
            if item in [
                "Delivery:", "Pick up:", "Telephone number:", "Fax number:",
                "Type / cab:", "VIN:", "Height:", "Weight:", "Latest delivery:",
                "Pick up until:", "Pick up untill:",
            ]:
                break

            # Kody lokalizacji typu AUSL16/I00780 nie sa czescia adresu.
            if not collected and re.match(r"^[A-Z]\w{3,8}$", item):
                continue

            collected.append(item)
            if len(collected) >= 3:
                return "\n".join(collected)

    return ""


def get_address_block(driver, label):
    return (
        address_block_near_label(driver, label)
        or address_block_after_label(driver, label)
        or address_block_from_text(driver, label)
    )


def collect_all_frame_texts(driver):
    chunks = []

    def walk(path, depth):
        try:
            text = driver.find_element(By.TAG_NAME, "body").text
            chunks.append((path, text))
        except Exception as exc:
            chunks.append((path, f"[brak body: {exc}]"))

        if depth <= 0:
            return

        frames = driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame")
        for idx, frame in enumerate(frames):
            try:
                driver.switch_to.frame(frame)
                walk(f"{path}/frame[{idx}]", depth - 1)
                driver.switch_to.parent_frame()
            except Exception as exc:
                chunks.append((f"{path}/frame[{idx}]", f"[blad ramki: {exc}]"))
                try:
                    driver.switch_to.parent_frame()
                except Exception:
                    driver.switch_to.default_content()

    driver.switch_to.default_content()
    walk("default", 4)
    driver.switch_to.default_content()
    return chunks


def save_debug_details(driver, index, log_cb):
    try:
        os.makedirs(DEBUG_DIR, exist_ok=True)
        path = os.path.join(DEBUG_DIR, f"details_{index:03d}.txt")
        chunks = collect_all_frame_texts(driver)

        with open(path, "w", encoding="utf-8") as handle:
            handle.write(f"URL: {driver.current_url}\n\n")
            for frame_path, text in chunks:
                handle.write(f"\n===== {frame_path} =====\n")
                handle.write(text)
                handle.write("\n")

        log_cb(f"Debug szczegolow zapisany: {path}")
    except Exception as exc:
        log_cb(f"Nie udalo sie zapisac debug szczegolow: {exc}")


def extract_phone_number(driver):
    try:
        els = driver.find_elements(
            By.XPATH,
            "//*[normalize-space(translate(., '\u00a0', ' '))='Telephone number:']/following::td[1]",
        )
        numbers = [clean_text(el.text) for el in els if clean_text(el.text)]

        if len(numbers) >= 2:
            return numbers[1]
        if numbers:
            return numbers[0]
    except Exception:
        pass

    return ""


def click_tab(driver, wait, label):
    try:
        el = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, f"//*[normalize-space(translate(., '\u00a0', ' '))='{label}']")
            )
        )
        el.click()
        time.sleep(0.3)
        return True
    except Exception:
        return False


def switch_to_frame_containing(driver, texts, max_depth=4):
    driver.switch_to.default_content()

    def search(depth):
        try:
            body = driver.find_element(By.TAG_NAME, "body")
            page_text = clean_text(body.text)
            if any(text in page_text for text in texts):
                return True
        except Exception:
            pass

        if depth <= 0:
            return False

        frames = driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame")

        for frame in frames:
            try:
                driver.switch_to.frame(frame)
                if search(depth - 1):
                    return True
                driver.switch_to.parent_frame()
            except Exception:
                try:
                    driver.switch_to.parent_frame()
                except Exception:
                    driver.switch_to.default_content()

        return False

    return search(max_depth)


def execute_json_script(driver, script):
    raw = driver.execute_script(script)
    if isinstance(raw, dict):
        return raw
    if not raw:
        return {}
    return json.loads(raw)


def execute_click_label(driver, label):
    try:
        return bool(driver.execute_script(CLICK_LABEL_JS, label))
    except Exception:
        return False


def collect_units_fast(driver):
    data = execute_json_script(driver, LIST_UNITS_JS)
    units = data.get("units") or []
    cleaned = []

    for unit in units:
        href = clean_text(unit.get("href", ""))
        vehicle_no = clean_text(unit.get("vehicleNo", ""))
        if not href or not vehicle_no:
            continue
        cleaned.append(
            {
                "url": href,
                "vehicle_no": vehicle_no,
                "to_no": clean_text(unit.get("toNo", "")),
            }
        )

    return cleaned


def wait_for_js_text(driver, texts, timeout=8):
    def has_text(_):
        try:
            data = execute_json_script(
                driver,
                """
                return (function() {
                  const chunks = [];
                  function collect(win) {
                    try {
                      chunks.push((win.document.body && win.document.body.innerText) || '');
                      for (let i = 0; i < win.frames.length; i++) collect(win.frames[i]);
                    } catch(e) {}
                  }
                  collect(window);
                  return JSON.stringify({text: chunks.join('\\n')});
                })();
                """,
            )
            page_text = data.get("text", "")
            return any(text in page_text for text in texts)
        except Exception:
            return False

    return WebDriverWait(driver, timeout).until(has_text)


def details_fast(driver):
    return execute_json_script(driver, DETAILS_EXTRACT_JS)


def address_dict_to_parts(address):
    address = address or {}
    return (
        clean_text(address.get("name", "")),
        clean_text(address.get("street", "")),
        clean_text(address.get("country", "")),
        clean_text(address.get("zipcode", "")),
        clean_text(address.get("city", "")),
    )


def wait_first_usable_css(driver, selectors, timeout=10):
    selector = ", ".join(selectors)

    def find_usable(_):
        elements = driver.find_elements(By.CSS_SELECTOR, selector)
        for element in elements:
            try:
                if element.is_displayed() and element.is_enabled():
                    return element
            except StaleElementReferenceException:
                continue
        return False

    return WebDriverWait(driver, timeout).until(find_usable)


def click_first_usable_css(driver, selectors, timeout=8):
    element = wait_first_usable_css(driver, selectors, timeout=timeout)
    try:
        element.click()
    except Exception:
        driver.execute_script("arguments[0].click();", element)
    return element


def click_optional_login_continue(driver):
    selectors = [
        "#idSIButton9",
        "input[type='submit']",
        "button[type='submit']",
        "button[id*='continue']",
        "button[id*='ok']",
    ]

    try:
        click_first_usable_css(driver, selectors, timeout=2)
        return True
    except Exception:
        return False


def scrape_man(login, password, log_cb, progress_cb):
    opts = Options()
    opts.page_load_strategy = "eager"
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option(
        "prefs",
        {
            "profile.managed_default_content_settings.images": 2,
            "profile.default_content_setting_values.notifications": 2,
        },
    )

    driver = None
    wait = None
    results = []

    try:
        log_cb("Uruchamiam przegladarke Chrome...")
        driver = webdriver.Chrome(options=opts)
        wait = WebDriverWait(driver, 15)

        log_cb("Otwieranie strony MAN MOTOS...")
        driver.get(LOGIN_URL)
        time.sleep(1)

        auto_logged = False
        try:
            login_field = wait_first_usable_css(
                driver,
                [
                "input[type='email']", "input[name='loginfmt']",
                "#i0116", "input[type='text']", "input[name='username']",
                ],
                timeout=18,
            )
            login_field.clear()
            login_field.send_keys(login)
            log_cb("Wpisano login...")

            click_first_usable_css(
                driver,
                ["input[type='submit']", "button[type='submit']", "#idSIButton9", "button[id*='next']"],
                timeout=4,
            )

            password_field = wait_first_usable_css(
                driver,
                ["input[type='password']", "input[name='passwd']", "#i0118", "input[name='password']"],
                timeout=18,
            )
            password_field.clear()
            password_field.send_keys(password)
            log_cb("Wpisano haslo...")

            click_first_usable_css(
                driver,
                ["input[type='submit']", "button[type='submit']", "#idSIButton9", "button[id*='login']"],
                timeout=4,
            )
            log_cb("Kliknieto przycisk logowania...")
            click_optional_login_continue(driver)

            auto_logged = True
        except Exception:
            pass

        if not auto_logged:
            log_cb("Automatyczne logowanie nie powiodlo sie.")
            log_cb("Zaloguj sie recznie w przegladarce, masz 3 minuty.")
        else:
            log_cb("Oczekiwanie na przekierowanie...")

        try:
            WebDriverWait(driver, 180).until(
                lambda d: "motos.man.eu" in d.current_url and "oauth" not in d.current_url
            )
            log_cb("Zalogowano. Ladowanie listy jednostek...")
        except TimeoutException:
            log_cb("Nie udalo sie zalogowac. Sprawdz login i haslo.")
            return []

        wait_for_js_text(driver, ["Vehic. no", "Pick up", "TO-No.", TARGET_CODE], timeout=20)
        log_cb(f"Szukam jednostek z kodem '{TARGET_CODE}' szybkim odczytem...")

        units = []
        page = 1

        while True:
            new = collect_units_fast(driver)
            existing_urls = {unit["url"] for unit in units}
            added = [unit for unit in new if unit["url"] not in existing_urls]
            units.extend(added)
            log_cb(f"Strona {page}: znaleziono nowych jednostek {len(added)}, razem {len(units)}.")

            if page > 1 and not added:
                break

            next_page = page + 1
            if not execute_click_label(driver, str(next_page)):
                break

            page = next_page
            time.sleep(0.35)

            try:
                wait_for_js_text(driver, ["Vehic. no", "Pick up", "TO-No."], timeout=5)
            except Exception:
                break

            if page > 20:
                break

        if not units:
            log_cb(f"Nie znaleziono jednostek z kodem '{TARGET_CODE}'.")
            return []

        log_cb(f"Znaleziono {len(units)} jednostek z kodem {TARGET_CODE}.")

        total = len(units)
        for i, unit in enumerate(units, 1):
            log_cb(f"Pobieram jednostke {i}/{total}...")
            progress_cb((i - 1) / total)

            try:
                driver.get(unit["url"])
                wait_for_js_text(driver, ["TO details", "Type / cab:", "VIN:", "Pick up:", "Delivery:"], timeout=15)

                if ENABLE_DEBUG_DETAILS and i == 1:
                    save_debug_details(driver, i, log_cb)

                execute_click_label(driver, "TO details")
                time.sleep(0.15)
                details = details_fast(driver)

                pickup = details.get("pickupAddress") or {}
                delivery = details.get("deliveryAddress") or {}
                _, pu_street, pu_country, pu_zipcode, pu_city = address_dict_to_parts(pickup)
                _, dl_street, dl_country, dl_zipcode, dl_city = address_dict_to_parts(delivery)

                product = clean_text(details.get("typeCab", ""))

                if is_bialezyce_100(pu_street) and execute_click_label(driver, "Vehicle details"):
                    time.sleep(0.15)
                    vehicle_details = details_fast(driver)
                    product = append_after_comma(product, vehicle_details.get("vehicleType", ""))

                row = {column: "" for column in COLUMNS}
                row["NAZWA / SKROT"] = "MAN"
                row["NUMER TELEFONU"] = clean_text(details.get("telephone", ""))
                row["ZALADUNEK PLANOWANA DATA"] = clean_text(details.get("pickupUntil", ""))
                row["ZALADUNEK KRAJ"] = pu_country
                row["ZALADUNEK KOD POCZTOWY"] = pu_zipcode
                row["ZALADUNEK MIASTO"] = pu_city
                row["ZALADUNEK ULICA NUMER"] = pu_street
                row["DOSTAWA PLANOWANA DATA"] = clean_text(details.get("latestDelivery", ""))
                row["DOSTAWA KRAJ"] = dl_country
                row["DOSTAWA KOD POCZTOWY"] = dl_zipcode
                row["DOSTAWA MIASTO"] = dl_city
                row["DOSTAWA ULICA NUMER"] = dl_street
                row["VIN NUMER REFERENCYJNY"] = clean_text(unit.get("vehicle_no", ""))
                row["TOWAR NAZWA URZADZENIA"] = product
                row["DLUGOSC [mm]"] = clean_text(details.get("totalLength", ""))
                row["WYSOKOSC [mm]"] = clean_text(details.get("height", ""))
                row["WAGA [kg]"] = clean_text(details.get("weight", ""))
                row["ROZSTAW OSI [mm]"] = clean_text(details.get("wheelOverhang", ""))
                row["BOOKING PRZESYLKA TMS"] = clean_text(unit.get("to_no", ""))
                row["WYSZUKIWANIE"] = clean_text(details.get("vin", ""))

                key_values = [
                    row["VIN NUMER REFERENCYJNY"],
                    row["WYSZUKIWANIE"],
                    row["ZALADUNEK PLANOWANA DATA"],
                    row["DOSTAWA PLANOWANA DATA"],
                    row["ZALADUNEK MIASTO"],
                    row["DOSTAWA MIASTO"],
                ]

                if any(clean_text(value) for value in key_values):
                    results.append(row)
                    log_cb(
                        "Dodano: "
                        f"Vehic.no={row['VIN NUMER REFERENCYJNY']}, "
                        f"VIN={row['WYSZUKIWANIE']}, "
                        f"zaladunek={row['ZALADUNEK PLANOWANA DATA']}, "
                        f"dostawa={row['DOSTAWA PLANOWANA DATA']}"
                    )
                else:
                    log_cb(f"Jednostka {i}: brak danych kluczowych, pomijam w eksporcie.")

                progress_cb(i / total)

            except Exception as e:
                log_cb(f"Blad przy jednostce {i}: {e}")

        log_cb("Pobieranie danych zakonczone!")

    except Exception as e:
        log_cb(f"Blad krytyczny: {e}")
        write_exception_log("Blad krytyczny scrapera", e)
    finally:
        if driver:
            driver.quit()

    return results


def export_xlsx(data, output_dir):
    data = [row for row in data if any(clean_text(value) for value in row.values())]
    if not data:
        raise ValueError("Brak niepustych danych do eksportu.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jednostki kod 20"

    hdr_fill = PatternFill("solid", fgColor="1C2B4A")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor="F0F4FA")
    bd = Side(style="thin", color="D0D7E3")
    brd = Border(left=bd, right=bd, top=bd, bottom=bd)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = brd
        cell.alignment = ctr

    ws.row_dimensions[1].height = 35

    for ri, row in enumerate(data, 2):
        fill = alt_fill if ri % 2 == 0 else None

        for ci, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.border = brd
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill

        ws.row_dimensions[ri].height = 20

    for ci in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

    ws.freeze_panes = "A2"

    filename = "MAN extract.xlsx"
    filepath = os.path.join(output_dir, filename)

    try:
        wb.save(filepath)
    except PermissionError as exc:
        raise FileLockedError(filepath) from exc

    return filepath


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.title(f"Ex MAN v{APP_VERSION}")
        try:
            self.iconbitmap(resource_path(ICON_FILE))
        except Exception:
            pass

        self.geometry("560x635")
        self.resizable(False, False)
        self.output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        self.saved_login = load_saved_login()

        hdr = ctk.CTkFrame(self, fg_color="transparent", corner_radius=0)
        hdr.pack(fill="x", padx=4, pady=(6, 0))

        try:
            self.banner_img = ctk.CTkImage(
                light_image=Image.open(resource_path(BANNER_FILE)),
                dark_image=Image.open(resource_path(BANNER_FILE)),
                size=(552, 104),
            )
            ctk.CTkLabel(hdr, text="", image=self.banner_img).pack()
        except Exception:
            fallback = ctk.CTkFrame(hdr, fg_color="#1C2B4A", corner_radius=0)
            fallback.pack(fill="x")

            ctk.CTkLabel(
                fallback,
                text=f"MAN MOTOS Ekstraktor v{APP_VERSION}",
                font=ctk.CTkFont(size=20, weight="bold"),
                text_color="white",
            ).pack(pady=18)

            ctk.CTkLabel(
                fallback,
                text="Pobiera jednostki z kodem 20 i eksportuje do pliku Excel",
                font=ctk.CTkFont(size=12),
                text_color="#A0B4D0",
            ).pack(pady=(0, 14))

        frm = ctk.CTkFrame(self, corner_radius=12)
        frm.pack(fill="x", padx=24, pady=18)

        ctk.CTkLabel(frm, text="Login (e-mail)", anchor="w").pack(fill="x", padx=16, pady=(14, 2))
        self.login_e = ctk.CTkEntry(frm, placeholder_text="twoj@email.com", height=38)
        self.login_e.pack(fill="x", padx=16, pady=(0, 10))
        if self.saved_login["email"]:
            self.login_e.insert(0, self.saved_login["email"])

        ctk.CTkLabel(frm, text="Haslo", anchor="w").pack(fill="x", padx=16, pady=(0, 2))
        self.pass_e = ctk.CTkEntry(frm, placeholder_text="********", show="*", height=38)
        self.pass_e.pack(fill="x", padx=16, pady=(0, 10))
        if self.saved_login["password"]:
            self.pass_e.insert(0, self.saved_login["password"])

        self.remember_login_var = ctk.BooleanVar(value=self.saved_login["remember"])
        self.remember_login_cb = ctk.CTkCheckBox(
            frm,
            text="Zapamietaj login i haslo na tym komputerze",
            variable=self.remember_login_var,
        )
        self.remember_login_cb.pack(anchor="w", padx=16, pady=(0, 12))

        ctk.CTkLabel(frm, text="Folder zapisu", anchor="w").pack(fill="x", padx=16, pady=(0, 2))
        save_row = ctk.CTkFrame(frm, fg_color="transparent")
        save_row.pack(fill="x", padx=16, pady=(0, 14))

        self.output_dir_var = ctk.StringVar(value=self.output_dir)
        self.output_dir_e = ctk.CTkEntry(save_row, textvariable=self.output_dir_var, height=38)
        self.output_dir_e.pack(side="left", fill="x", expand=True, padx=(0, 8))

        self.output_dir_btn = ctk.CTkButton(
            save_row,
            text="Wybierz...",
            width=105,
            height=38,
            command=self.choose_output_dir,
        )
        self.output_dir_btn.pack(side="right")

        self.btn = ctk.CTkButton(
            self,
            text="Pobierz dane i zapisz do Excel",
            height=46,
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color="#E8511A",
            hover_color="#C7400E",
            command=self.start,
        )
        self.btn.pack(fill="x", padx=24, pady=(0, 12))

        self.progress = ctk.CTkProgressBar(self, height=10)
        self.progress.set(0)
        self.progress.pack(fill="x", padx=24, pady=(0, 8))

        self.status_label = ctk.CTkLabel(
            self,
            text="Gotowe do pracy",
            anchor="w",
            height=34,
            fg_color="#F0F4FA",
            text_color="#1C2B4A",
            corner_radius=8,
            font=ctk.CTkFont(size=12),
        )
        self.status_label.pack(fill="x", padx=24, pady=(0, 20))

    def log(self, msg):
        self.after(0, self._log_safe, msg)

    def _log_safe(self, msg):
        write_app_log(msg)
        self.status_label.configure(text=msg)

    def set_progress(self, val, msg=None):
        def update():
            self.progress.set(val)
            if msg:
                self.status_label.configure(text=msg)

        self.after(0, update)

    def set_button_state(self, state, text):
        self.after(0, lambda: self.btn.configure(state=state, text=text))

    def ask_retry_locked_file(self, filepath):
        answer = {"retry": False}
        done = threading.Event()

        def ask():
            answer["retry"] = messagebox.askretrycancel(
                "Plik Excel jest otwarty",
                "Nie moge zapisac pliku, poniewaz jest otwarty w Excelu.\n\n"
                f"Zamknij plik:\n{filepath}\n\n"
                "Nastepnie kliknij 'Ponow', aby zapisac ponownie.",
                parent=self,
            )
            done.set()

        self.after(0, ask)
        done.wait()
        return answer["retry"]

    def export_with_retry(self, data, output_dir):
        while True:
            try:
                return export_xlsx(data, output_dir)
            except FileLockedError as e:
                self.log("Plik MAN extract.xlsx jest otwarty. Zamknij go w Excelu.")
                if not self.ask_retry_locked_file(e.filepath):
                    self.log("Zapis anulowany. Dane nie zostaly zapisane.")
                    return ""

    def choose_output_dir(self):
        selected = filedialog.askdirectory(
            title="Wybierz folder zapisu pliku Excel",
            initialdir=self.output_dir,
        )

        if selected:
            self.output_dir = selected
            self.output_dir_var.set(selected)

    def start(self):
        login = self.login_e.get().strip()
        password = self.pass_e.get().strip()
        output_dir = self.output_dir_var.get().strip()

        if not login or not password:
            self.log("Podaj login i haslo")
            return

        if not output_dir or not os.path.isdir(output_dir):
            self.log("Wybierz poprawny folder zapisu")
            return

        if self.remember_login_var.get():
            try:
                save_login(login, password)
            except Exception as e:
                self.log(f"Nie udalo sie zapisac danych logowania: {e}")
        else:
            delete_saved_login()

        self.btn.configure(state="disabled", text="Trwa pobieranie...")
        self.progress.set(0)
        self.status_label.configure(text="Uruchamiam pobieranie...")

        def run():
            try:
                data = scrape_man(login, password, self.log, self.set_progress)

                if data:
                    try:
                        filepath = self.export_with_retry(data, output_dir)
                        if filepath:
                            self.set_progress(1, f"Gotowe. Zapisano: {filepath}")
                    except ValueError as e:
                        self.log(str(e))
                    except Exception as e:
                        self.log(f"Blad eksportu: {e}")
                        write_exception_log("Blad eksportu", e)
                else:
                    self.log("Brak danych do zapisania")
            except Exception as e:
                self.log(f"Blad programu: {e}")
                self.log(f"Szczegoly zapisano w: {LOG_FILE}")
                write_exception_log("Blad watku programu", e)
            finally:
                self.set_button_state("normal", "Pobierz dane i zapisz do Excel")


        threading.Thread(target=run, daemon=True).start()


if __name__ == "__main__":
    try:
        app = App()
        app.mainloop()
    except Exception as exc:
        print("BLAD PROGRAMU:")
        print(exc)
        input("Nacisnij Enter, aby zamknac...")
